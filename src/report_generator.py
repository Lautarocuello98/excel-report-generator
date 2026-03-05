from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def generate_excel_report(
    report_path: Path,
    df_clean: pd.DataFrame,
    kpis: dict,
    currency: str,
    sources: Iterable[str],
    chart_files: list[Path] | None = None,
) -> None:
    """
    Creates an Excel report with:
    - Summary sheet (KPIs + sources)
    - Cleaned Data sheet (data)
    - Optional charts embedded
    """
    report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"

    title = f"Sales Report ({currency})"
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    rows = [
        ("Total Orders", kpis.get("total_orders", 0)),
        ("Total Units", round(float(kpis.get("total_units", 0.0)), 2)),
        ("Total Revenue", round(float(kpis.get("total_revenue", 0.0)), 2)),
        ("Total Cost", round(float(kpis.get("total_cost", 0.0)), 2)),
        ("Total Profit", round(float(kpis.get("total_profit", 0.0)), 2)),
        ("Avg Order Value", round(float(kpis.get("avg_order_value", 0.0)), 2)),
    ]

    start_row = 3
    for i, (k, v) in enumerate(rows, start=start_row):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = v

    ws["A10"] = "Sources"
    ws["A10"].font = Font(bold=True)
    for idx, s in enumerate(list(sources), start=11):
        ws[f"A{idx}"] = s

    _set_col_widths(ws, {"A": 28, "B": 18})

    # Embed charts if provided
    if chart_files:
        anchor_row = 3
        anchor_col = 4  # column D
        for i, img_path in enumerate(chart_files):
            try:
                img = XLImage(str(img_path))
                img.anchor = ws.cell(row=anchor_row + i * 18, column=anchor_col).coordinate
                ws.add_image(img)
            except Exception:
                # If image embedding fails, ignore (report still valid)
                pass

    # --- Cleaned Data ---
    ws2 = wb.create_sheet("Cleaned Data")

    # If processor already added calculated columns, include them if present
    df_out = df_clean.copy()
    if {"quantity", "unit_price", "unit_cost"}.issubset(df_out.columns):
        if "revenue" not in df_out.columns:
            df_out["revenue"] = df_out["quantity"] * df_out["unit_price"]
        if "cost" not in df_out.columns:
            df_out["cost"] = df_out["quantity"] * df_out["unit_cost"]
        if "profit" not in df_out.columns:
            df_out["profit"] = df_out["revenue"] - df_out["cost"]

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws2.append(r)

    ws2.freeze_panes = "A2"
    _set_col_widths(ws2, {"A": 18, "B": 18, "C": 28, "D": 12, "E": 12, "F": 12})

    # --- Top Products ---
    top = kpis.get("top_products")
    if isinstance(top, pd.DataFrame) and not top.empty:
        ws3 = wb.create_sheet("Top Products")
        for r in dataframe_to_rows(top, index=False, header=True):
            ws3.append(r)
        ws3.freeze_panes = "A2"
        _set_col_widths(ws3, {"A": 32, "B": 16})

    wb.save(report_path)